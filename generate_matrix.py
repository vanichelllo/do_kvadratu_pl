import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Ваші налаштування
FOLDER_PATH = 'import_files'  # Папка з вашими PDF
EXCEL_PATH = 'matrix.xlsx'
FREE_TOPICS = [1, 3, 4, 6, 7, 11, 17, 19, 20, 29, 30, 31, 35, 37, 38]


def generate():
    wb = Workbook()
    ws = wb.active
    ws.title = "Матриця конспектів"

    # Заголовки (додано колонку для пакетів)
    headers = ["Назва матеріалу", "Ціна", "Файл", "Безкоштовно (1/0)", "Пакет (1/0)"]
    ws.append(headers)

    # Стиль заголовків
    for col in range(1, 6):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2C3E50", fill_type="solid")
        ws.column_dimensions[cell.column_letter].width = 25
    ws.column_dimensions['A'].width = 50

    # Читаємо всі файли з вашої папки
    if not os.path.exists(FOLDER_PATH):
        print(f"Помилка: Папку '{FOLDER_PATH}' не знайдено!")
        return

    files = sorted([f for f in os.listdir(FOLDER_PATH) if f.endswith('.pdf')])

    for filename in files:
        # Шукаємо номер на початку файлу (напр. "1. Числові множини.pdf")
        match = re.match(r'^(\d+)\.\s*(.+)\.pdf$', filename)

        if match:
            num = int(match.group(1))
            title = f"{num}. {match.group(2)}"

            # Застосовуємо вашу логіку цін
            if num in FREE_TOPICS:
                price = 0
                is_free = 1
            else:
                price = 19
                is_free = 0

            ws.append([title, price, filename, is_free, 0])
        else:
            # Якщо файл не має цифри на початку
            ws.append([filename.replace('.pdf', ''), 19, filename, 0, 0])

    # Додаємо пакети в кінець таблиці
    ws.append(["Вся алгебра для НМТ", 399, "Bundle_01.pdf", 0, 1])
    ws.append(["Вся геометрія для НМТ", 249, "Bundle_02.pdf", 0, 1])
    ws.append(["Весь курс НМТ", 599, "Bundle_03.pdf", 0, 1])

    wb.save(EXCEL_PATH)
    print(f"Супер! Файл {EXCEL_PATH} успішно створено. Додано {len(files)} тем та 3 пакети.")


if __name__ == '__main__':
    generate()