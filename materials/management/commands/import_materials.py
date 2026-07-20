import os
from django.core.management.base import BaseCommand
from django.core.files import File
from openpyxl import load_workbook
from materials.models import StudyMaterial, Category


class Command(BaseCommand):
    help = 'Розумний імпорт з новими правилами цін, категорій та збереженням PDF'

    def handle(self, *args, **kwargs):
        excel_path = 'matrix.xlsx'
        pdf_folder = 'import_files'

        # 1. Створюємо категорію (система перевірить, як названо поле у вашій базі: name чи title)
        try:
            category, _ = Category.objects.get_or_create(name="Підготовка до НМТ")
        except:
            category, _ = Category.objects.get_or_create(title="Підготовка до НМТ")

        wb = load_workbook(excel_path)
        ws = wb.active

        self.stdout.write("Починаємо оновлення бази даних...")

        for row in ws.iter_rows(min_row=2, values_only=True):
            excel_title, original_price, filename, is_free, is_bundle = row

            if not excel_title or not filename:
                continue

            # 2. Формуємо нові назви та ціни за вашими правилами
            new_title = excel_title
            price = 25  # Всі окремі теми по 25 грн

            # Правила для пакетів
            if is_bundle:
                price = original_price  # Пакети зберігають свою стару ціну з Excel
                if "Пакет 1" in excel_title:
                    new_title = "Вся алгебра для НМТ"
                elif "Пакет 2" in excel_title:
                    new_title = "Вся геометрія для НМТ"
                elif "Пакет 3" in excel_title:
                    new_title = "Весь курс для НМТ"

            # 3. Шукаємо матеріал у базі (за старою або новою назвою)
            material = StudyMaterial.objects.filter(title=excel_title).first()
            if not material:
                material = StudyMaterial.objects.filter(title=new_title).first()

            created = False
            if not material:
                material = StudyMaterial(title=new_title)
                created = True

            # 4. Оновлюємо текст, ціни та прив'язуємо категорію
            material.title = new_title
            material.price = price
            material.is_free = bool(is_free)
            material.is_bundle = bool(is_bundle)
            material.is_published = True
            material.category = category
            material.save()

            # 5. ФАЙЛИ БІЛЬШЕ НЕ ПЕРЕЗАПИСУЮТЬСЯ!
            # Скрипт приклеїть і завантажить PDF тільки якщо це новий запис,
            # або якщо файл випадково порожній.
            if created or not material.file:
                pdf_path = os.path.join(pdf_folder, filename)
                if os.path.exists(pdf_path):
                    with open(pdf_path, 'rb') as f:
                        material.file.save(filename, File(f), save=True)
                    self.stdout.write(self.style.SUCCESS(f'Створено з PDF: {new_title}'))
            else:
                self.stdout.write(f'Оновлено текст і ціну: {new_title} (PDF не чіпали)')

        self.stdout.write(self.style.SUCCESS('Готово! Базу оновлено, ваші старі файли у безпеці.'))